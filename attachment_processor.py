#!/usr/bin/env python3
"""Email Attachment Processor — extract text from common file types.

Supports: PDF, DOCX, TXT, CSV, HTML, images (OCR), XLS/XLSX
Extracted text feeds into the NLP training pipeline.
Attachments stored compressed + encrypted alongside email body.
"""

import io
import os
import sys
import logging
import tempfile
from pathlib import Path

sys.path.insert(0, "/home/aielevate")

log = logging.getLogger("attachment-processor")


def extract_text(filename, content_bytes):
    """Extract text from an attachment by file type.

    Args:
        filename: Original filename (used to detect type)
        content_bytes: Raw file bytes

    Returns:
        Extracted text string, or empty string if unsupported/failed
    """
    ext = Path(filename).suffix.lower()

    try:
        if ext == ".txt":
            return _extract_txt(content_bytes)
        elif ext == ".pdf":
            return _extract_pdf(content_bytes)
        elif ext in (".doc", ".docx"):
            return _extract_docx(content_bytes)
        elif ext == ".csv":
            return _extract_csv(content_bytes)
        elif ext in (".xls", ".xlsx"):
            return _extract_excel(content_bytes)
        elif ext in (".html", ".htm"):
            return _extract_html(content_bytes)
        elif ext == ".svg":
            return _extract_svg(content_bytes)
        elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif", ".webp", ".ico"):
            return _extract_image_ocr(content_bytes)
        elif ext in (".odt",):
            return _extract_odt(content_bytes)
        elif ext in (".ods",):
            return _extract_ods(content_bytes)
        elif ext in (".odp",):
            return _extract_odp(content_bytes)
        elif ext in (".odg", ".odf"):
            return _extract_odf_generic(content_bytes)
        elif ext in (".rtf",):
            return _extract_rtf(content_bytes)
        elif ext in (".json",):
            return _extract_txt(content_bytes)
        elif ext in (".md", ".rst", ".log", ".ini", ".cfg", ".yaml", ".yml", ".toml", ".xml"):
            return _extract_txt(content_bytes)
        elif ext in (".eml", ".msg"):
            return _extract_txt(content_bytes)
        elif ext in (".ppt", ".pptx"):
            return _extract_pptx(content_bytes)
        else:
            log.debug(f"Unsupported attachment type: {ext} ({filename})")
            return ""
    except Exception as e:
        log.warning(f"Failed to extract text from {filename}: {e}")
        return ""


def _extract_txt(content_bytes):
    """Plain text extraction."""
    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            return content_bytes.decode(encoding)[:50000]
        except (UnicodeDecodeError, AttributeError):
            continue
    return ""


def _extract_pdf(content_bytes):
    """Extract text from PDF."""
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(content_bytes))
        text = ""
        for page in reader.pages[:50]:  # Limit to 50 pages
            text += page.extract_text() or ""
        return text[:50000]
    except ImportError:
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content_bytes)) as pdf:
                text = ""
                for page in pdf.pages[:50]:
                    text += page.extract_text() or ""
                return text[:50000]
        except ImportError:
            log.warning("No PDF library available (install PyPDF2 or pdfplumber)")
            return ""


def _extract_docx(content_bytes):
    """Extract text from DOCX."""
    try:
        import docx
        doc = docx.Document(io.BytesIO(content_bytes))
        text = "\n".join(para.text for para in doc.paragraphs)
        return text[:50000]
    except ImportError:
        log.warning("python-docx not installed")
        return ""


def _extract_csv(content_bytes):
    """Extract text from CSV — headers + first 100 rows."""
    import csv
    text = _extract_txt(content_bytes)
    if text:
        lines = text.split("\n")[:101]  # Header + 100 rows
        return "\n".join(lines)
    return ""


def _extract_excel(content_bytes):
    """Extract text from Excel files."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True)
        text = ""
        for sheet in wb.sheetnames[:5]:  # First 5 sheets
            ws = wb[sheet]
            for row in ws.iter_rows(max_row=100, values_only=True):
                text += " ".join(str(cell) for cell in row if cell) + "\n"
        return text[:50000]
    except ImportError:
        log.warning("openpyxl not installed")
        return ""


def _extract_html(content_bytes):
    """Extract text from HTML."""
    from html.parser import HTMLParser

    class TextExtractor(HTMLParser):
        def __init__(self):
            super().__init__()
            self.text = []
            self._skip = False

        def handle_starttag(self, tag, attrs):
            if tag in ("script", "style"):
                self._skip = True

        def handle_endtag(self, tag):
            if tag in ("script", "style"):
                self._skip = False

        def handle_data(self, data):
            if not self._skip:
                self.text.append(data.strip())

    parser = TextExtractor()
    parser.feed(_extract_txt(content_bytes))
    return " ".join(t for t in parser.text if t)[:50000]


def _extract_svg(content_bytes):
    """Extract text from SVG (XML-based vector format)."""
    import re
    try:
        svg = content_bytes.decode("utf-8", errors="replace")
        # Extract text elements
        texts = re.findall(r"<text[^>]*>([^<]+)</text>", svg)
        # Also extract tspan content
        texts += re.findall(r"<tspan[^>]*>([^<]+)</tspan>", svg)
        # Extract title and desc
        texts += re.findall(r"<title>([^<]+)</title>", svg)
        texts += re.findall(r"<desc>([^<]+)</desc>", svg)
        return " ".join(texts).strip()[:20000]
    except Exception as e:
        log.debug(f"SVG extraction failed: {e}")
        return ""


def _extract_image_ocr(content_bytes):
    """Extract text from images using OCR + vision model for visual understanding.

    Two passes:
    1. OCR (Tesseract) — extracts visible text from the image
    2. Vision model (Ollama/moondream) — describes what the image shows
    Combined gives both textual and visual understanding.
    """
    text_parts = []

    # Pass 1: OCR for visible text
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(io.BytesIO(content_bytes))
        ocr_text = pytesseract.image_to_string(img).strip()
        if ocr_text and len(ocr_text) > 10:
            text_parts.append(f"[OCR text]: {ocr_text}")
    except ImportError:
        log.debug("pytesseract or PIL not installed — skipping OCR")
    except Exception as e:
        log.debug(f"OCR failed: {e}")

    # Pass 2: Vision model for image understanding
    try:
        description = _vision_describe(content_bytes)
        if description:
            text_parts.append(f"[Image description]: {description}")
    except Exception as e:
        log.debug(f"Vision analysis failed: {e}")

    return "\n".join(text_parts)[:20000]


def _vision_describe(content_bytes, prompt="Describe this image in detail. What does it show? Include any text, diagrams, charts, UI elements, or data visible."):
    """Describe an image using a local vision model (Ollama/moondream).

    Runs locally — no API cost, ~2-5 seconds per image.
    Falls back to Claude Vision API if local model unavailable.
    """
    import base64
    import json
    import urllib.request

    img_b64 = base64.b64encode(content_bytes).decode("utf-8")

    # Try local Ollama first (free, fast)
    try:
        payload = json.dumps({
            "model": "moondream",
            "prompt": prompt,
            "images": [img_b64],
            "stream": False,
        }).encode("utf-8")

        req = urllib.request.Request(
            "http://localhost:11434/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        resp = urllib.request.urlopen(req, timeout=90)
        result = json.loads(resp.read().decode("utf-8"))
        description = result.get("response", "").strip()
        if description:
            log.info(f"Vision (moondream): {len(description)} chars")
            return description[:5000]
    except Exception as e:
        log.debug(f"Ollama vision failed: {e}")

    # Fallback: Claude Vision API (costs ~$0.01-0.05 per image)
    try:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            key_file = Path("/opt/ai-elevate/credentials/anthropic-api-key.txt")
            if key_file.exists():
                api_key = key_file.read_text().strip()

        if api_key:
            from PIL import Image
            img = Image.open(io.BytesIO(content_bytes))
            fmt = img.format or "png"
            media_type = f"image/{fmt.lower()}"

            payload = json.dumps({
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 500,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                        {"type": "text", "text": prompt}
                    ]
                }]
            }).encode("utf-8")

            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                },
                method="POST"
            )
            resp = urllib.request.urlopen(req, timeout=30)
            result = json.loads(resp.read().decode("utf-8"))
            text = result.get("content", [{}])[0].get("text", "")
            if text:
                log.info(f"Vision (claude): {len(text)} chars")
                return text[:5000]
    except Exception as e:
        log.debug(f"Claude vision failed: {e}")

    return ""



def _extract_odt(content_bytes):
    """Extract text from OpenDocument Text (.odt)."""
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(content_bytes)) as z:
            if "content.xml" in z.namelist():
                xml = z.read("content.xml").decode("utf-8")
                # Strip XML tags
                import re
                text = re.sub(r"<[^>]+>", " ", xml)
                text = re.sub(r"\s+", " ", text).strip()
                return text[:50000]
    except Exception as e:
        log.warning(f"ODT extraction failed: {e}")
    return ""


def _extract_ods(content_bytes):
    """Extract text from OpenDocument Spreadsheet (.ods)."""
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(content_bytes)) as z:
            if "content.xml" in z.namelist():
                xml = z.read("content.xml").decode("utf-8")
                import re
                text = re.sub(r"<[^>]+>", " ", xml)
                text = re.sub(r"\s+", " ", text).strip()
                return text[:50000]
    except Exception as e:
        log.warning(f"ODS extraction failed: {e}")
    return ""


def _extract_odp(content_bytes):
    """Extract text from OpenDocument Presentation (.odp)."""
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(content_bytes)) as z:
            if "content.xml" in z.namelist():
                xml = z.read("content.xml").decode("utf-8")
                import re
                text = re.sub(r"<[^>]+>", " ", xml)
                text = re.sub(r"\s+", " ", text).strip()
                return text[:50000]
    except Exception as e:
        log.warning(f"ODP extraction failed: {e}")
    return ""


def _extract_odf_generic(content_bytes):
    """Extract text from any ODF format (.odg, .odf) via content.xml."""
    return _extract_odt(content_bytes)  # Same ZIP + XML structure


def _extract_rtf(content_bytes):
    """Extract text from Rich Text Format (.rtf)."""
    import re
    text = content_bytes.decode("latin-1", errors="replace")
    # Strip RTF control words
    text = re.sub(r"\\[a-z]+\d*\s?", " ", text)
    text = re.sub(r"[{}]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:50000]


def _extract_pptx(content_bytes):
    """Extract text from PowerPoint (.pptx)."""
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(content_bytes)) as z:
            text = ""
            for name in sorted(z.namelist()):
                if name.startswith("ppt/slides/slide") and name.endswith(".xml"):
                    xml = z.read(name).decode("utf-8")
                    import re
                    slide_text = re.sub(r"<[^>]+>", " ", xml)
                    slide_text = re.sub(r"\s+", " ", slide_text).strip()
                    text += slide_text + "\n"
            return text[:50000]
    except Exception as e:
        log.warning(f"PPTX extraction failed: {e}")
    return ""




def warmup_vision():
    """Pre-load the vision model so first request isn't slow.
    Call this once at gateway startup."""
    import json, urllib.request, base64, threading
    def _warmup():
        try:
            # Tiny 1x1 PNG
            payload = json.dumps({
                "model": "moondream",
                "prompt": "hi",
                "images": ["iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8/5+hHgAHggJ/PchI7wAAAABJRU5ErkJggg=="],
                "stream": False,
            }).encode()
            req = urllib.request.Request("http://localhost:11434/api/generate", data=payload, headers={"Content-Type": "application/json"})
            urllib.request.urlopen(req, timeout=120)
            log.info("Vision model warmed up")
        except Exception as e:
            log.debug(f"Vision warmup: {e}")
    threading.Thread(target=_warmup, daemon=True).start()

def process_attachments(files):
    """Process a list of uploaded files, return combined extracted text.

    Args:
        files: List of UploadFile objects from FastAPI

    Returns:
        dict with keys:
            text: Combined extracted text from all attachments
            filenames: List of processed filenames
            sizes: List of file sizes
            count: Number of attachments
    """
    if not files:
        return {"text": "", "filenames": [], "sizes": [], "count": 0}

    combined_text = []
    filenames = []
    sizes = []

    for f in files:
        try:
            if hasattr(f, "read"):
                content = f.read()
                if hasattr(f, "seek"):
                    f.seek(0)  # Reset for potential re-read
            elif hasattr(f, "file"):
                content = f.file.read()
                f.file.seek(0)
            elif isinstance(f, bytes):
                content = f
            else:
                continue

            fname = getattr(f, "filename", "unknown")
            filenames.append(fname)
            sizes.append(len(content))

            text = extract_text(fname, content)
            if text:
                combined_text.append(f"[Attachment: {fname}]\n{text}")
                log.info(f"Extracted {len(text)} chars from {fname}")

        except Exception as e:
            log.warning(f"Failed to process attachment: {e}")

    return {
        "text": "\n\n".join(combined_text),
        "filenames": filenames,
        "sizes": sizes,
        "count": len(filenames),
    }


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Attachment Processor")
    parser.add_argument("--file", type=str, help="Test extraction on a file")
    args = parser.parse_args()

    if args.file:
        p = Path(args.file)
        if p.exists():
            text = extract_text(p.name, p.read_bytes())
            print(f"Extracted {len(text)} chars from {p.name}")
            print(text[:500])
        else:
            print(f"File not found: {args.file}")
